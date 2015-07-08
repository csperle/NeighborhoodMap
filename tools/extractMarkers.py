# coding=utf8

import openpyxl
import unicodecsv
import re
import time

wb = openpyxl.load_workbook('../data/Vernetzung.xlsx')
sheet = wb.get_sheet_by_name('Vernetzung ueberarbeitet')
csvFile = open('../data/vernetzung.csv', 'w')
csvWriter = unicodecsv.writer(csvFile, delimiter=';', encoding='utf-8')


def processTags(tags):
    ptags = u""
    for val in re.split("[&,]", tags):
        stag = val.replace("/", "").strip()
        if len(stag) > 0:
            ptags += stag + ','
    return ptags[:-1]


def concTags(tag1, tag2, tag3, tag4):
    ctags = u""
    if isinstance(tag1, unicode):
        ctags += tag1 + "&"
    if isinstance(tag2, unicode):
        ctags += tag2 + "&"
    if isinstance(tag3, unicode):
        ctags += tag3 + "&"
    if isinstance(tag4, unicode):
        ctags += tag4 + "&"
    return ctags


for rowNum in range(1, sheet.get_highest_row() + 1):
    rowData = []
    anbieter = sheet.cell(row=rowNum, column=0).value
    tag1 = sheet.cell(row=rowNum, column=1).value
    tag2 = sheet.cell(row=rowNum, column=2).value
    tag3 = sheet.cell(row=rowNum, column=3).value
    tag4 = sheet.cell(row=rowNum, column=4).value
    beschreibung = sheet.cell(row=rowNum, column=5).value
    bemerkung = sheet.cell(row=rowNum, column=6).value
    adresse = sheet.cell(row=rowNum, column=7).value
    lat = sheet.cell(row=rowNum, column=8).value
    lng = sheet.cell(row=rowNum, column=9).value
    kontaktadresse = sheet.cell(row=rowNum, column=0).value
    kontakt = sheet.cell(row=rowNum, column=11).value
    ansprechperson = sheet.cell(row=rowNum, column=12).value
    email = sheet.cell(row=rowNum, column=13).value
    webseite = sheet.cell(row=rowNum, column=14).value
    kosten = sheet.cell(row=rowNum, column=15).value

    icon = "default"
    titel = ""
    tags = processTags(concTags(tag1, tag2, tag3, tag4))
    inserted = time.strftime("%Y-%m-%d %H:%M:%S")
    lastUpdate = ""

    rowData.append(rowNum)
    rowData.append(lat)
    rowData.append(lng)
    rowData.append(icon)
    rowData.append(anbieter)
    rowData.append(titel)
    rowData.append(beschreibung)
    rowData.append(tags)
    rowData.append(bemerkung)
    rowData.append(adresse)
    rowData.append(kontakt)
    rowData.append(ansprechperson)
    rowData.append(email)
    rowData.append(webseite)
    rowData.append(inserted)
    rowData.append(lastUpdate)

    # for colNum in range(0, 16):
        # print(rowNum, colNum)
        # print(sheet.cell(row=rowNum, column=colNum).value)
        # cellValue = sheet.cell(row=rowNum, column=colNum).value
        # print(type(cellValue))
        # if isinstance(cellValue, unicode):
            # print(cellValue)
            # rowData.append(processTags(cellValue))
        # else:
            # rowData.append(cellValue)
    csvWriter.writerow(rowData)
csvFile.close()
