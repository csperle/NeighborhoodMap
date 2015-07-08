# coding=utf8

import openpyxl
import unicodecsv
import re
import time

wb = openpyxl.load_workbook('../data/Vernetzung.xlsx')
sheet = wb.get_sheet_by_name('Vernetzung ueberarbeitet')
csvFile = open('../data/tags.csv', 'w')
csvWriter = unicodecsv.writer(csvFile, delimiter=';', encoding='utf-8')


def processTags(tags):
    ptags = u""
    for val in re.split("[&,]", tags):
        stag = val.replace("/", "").strip()
        if len(stag) > 0:
            ptags += stag + ','
    return ptags[:-1]


def concTags(tag1, tag2, tag3, tag4):
    ctags = u""
    if isinstance(tag1, unicode):
        ctags += tag1 + "&"
    if isinstance(tag2, unicode):
        ctags += tag2 + "&"
    if isinstance(tag3, unicode):
        ctags += tag3 + "&"
    if isinstance(tag4, unicode):
        ctags += tag4 + "&"
    return ctags

allTags = []
for rowNum in range(1, sheet.get_highest_row() + 1):
    tag1 = sheet.cell(row=rowNum, column=1).value
    tag2 = sheet.cell(row=rowNum, column=2).value
    tag3 = sheet.cell(row=rowNum, column=3).value
    tag4 = sheet.cell(row=rowNum, column=4).value
    tags = processTags(concTags(tag1, tag2, tag3, tag4))

    for tag in re.split("[,]", tags):
        if tag not in allTags:
            allTags.append(tag)

for i in range(0, len(allTags) - 1):
    rowData = [i + 1, allTags[i]]
    csvWriter.writerow(rowData)
csvFile.close()
